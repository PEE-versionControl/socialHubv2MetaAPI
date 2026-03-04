#!/usr/bin/env python3
"""
Excel Report Generator
======================
Generates .xlsx files matching the client's post-buy report template.

Layout per sheet:
  - Campaign header  (name, code, launch date, placement)
  - Post image       (fetched via Meta Graph API — no scraping)
  - Post Insight table  (organic metrics)
  - Ad Boosting Report  (paid metrics)

Usage:
    from excel_report import generate_excel_report
    buf = generate_excel_report(result)          # returns BytesIO
    buf = generate_excel_report(result, fb_token=..., ig_token=...)
"""

import io
import os
import re
import requests
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

try:
    from PIL import Image as _PILImage
except ImportError:
    _PILImage = None

load_dotenv()

# ── API constants ─────────────────────────────────────────────────
_BASE_URL = "https://graph.facebook.com/v24.0"

# ── Template colour palette ───────────────────────────────────────
_ORANGE   = "ED7D31"
_WHITE    = "FFFFFF"
_BLACK    = "000000"
_BLUE     = "0563C1"
_GREY_TXT = "757070"
_LIGHT_BG = "F2F2F2"

# ── Shared style builders ─────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=_BLACK, size=14, underline=False, italic=False) -> Font:
    return Font(
        name="Calibri", size=size, bold=bold, color=color,
        underline="single" if underline else None, italic=italic,
    )

def _border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


# ── Column widths (A=1 … K=11) matching template ─────────────────
_COL_WIDTHS = {
    1: 6,
    2: 43,   # ≈ 8 cm — wide enough for ad set names
    3: 52.71,
    4: 24.43,
    5: 25.71,
    6: 18,
    7: 18,
    8: 18,
    9: 18,
    10: 21.29,
    11: 18,
    12: 18,   # Post shares
    13: 18,   # Post saves
}


# ── Campaign header parsing ───────────────────────────────────────

def parse_campaign_header(adset_names: list, campaign_names: list = None) -> dict:
    """Extract code, campaign display name, and full campaign_name from ad data.

    Campaign name format: "(Reach)(type)#12345CampaignName"
    Adset name format:    "FB(Reach) (Bank) #12345 CampaignName"

    Returns {
        "code": "12345",
        "name": "CampaignName",
        "full_campaign_name": "(Reach)(type)#12345CampaignName"  # for filename
    }
    """
    def _extract(text):
        text = str(text).strip()
        if not text or text == "N/A":
            return None
        # Match #code with optional space before campaign name
        m = re.search(r'#(\d+)\s*(.*)', text)
        if m:
            return {
                "code": m.group(1),
                "name": m.group(2).strip() if m.group(2).strip() else text,
                "full_campaign_name": text,
            }
        return None

    # Try campaign_names first (most authoritative — direct from Ad Manager)
    if campaign_names:
        for cname in campaign_names:
            result = _extract(cname)
            if result:
                return result

    # Fallback: try adset_names
    for name in adset_names:
        result = _extract(name)
        if result:
            return result

    # Last fallback: use campaign_name as-is without code
    if campaign_names:
        for cname in campaign_names:
            cname = str(cname).strip()
            if cname and cname != "N/A":
                return {"code": "", "name": cname, "full_campaign_name": cname}

    return {"code": "", "name": "", "full_campaign_name": ""}


def format_date_range(date_range_str: str) -> str:
    """Convert "2026-01-08 to 2026-01-12" → "8 Jan - 12 Jan 2026 (5 Days)"."""
    if not date_range_str or " to " not in date_range_str:
        return date_range_str or ""
    try:
        parts = date_range_str.split(" to ")
        start = datetime.strptime(parts[0].strip(), "%Y-%m-%d")
        end   = datetime.strptime(parts[1].strip(), "%Y-%m-%d")
        days  = (end - start).days + 1
        return (
            f"{start.day} {start.strftime('%b')} - "
            f"{end.day} {end.strftime('%b')} {end.year} ({days} Days)"
        )
    except Exception:
        return date_range_str


def _get_ad_start_date(ad_metrics: list) -> str:
    """Extract the earliest ad start date from ad metrics.

    Checks date_start (from insights) and adset_start_time (from adset).
    Returns formatted string like "8 Jan 2026" or "" if not available.
    """
    earliest = None
    for am in ad_metrics:
        for field in ["date_start", "adset_start_time"]:
            raw = am.get(field, "")
            if not raw:
                continue
            try:
                # Handle ISO format "2026-01-08T..." or "2026-01-08"
                dt = datetime.fromisoformat(raw.replace("Z", "+00:00").replace("+0000", "+00:00"))
                if earliest is None or dt < earliest:
                    earliest = dt
            except (ValueError, TypeError):
                try:
                    dt = datetime.strptime(raw[:10], "%Y-%m-%d")
                    if earliest is None or dt < earliest:
                        earliest = dt
                except (ValueError, TypeError):
                    pass
    if earliest:
        return f"{earliest.day} {earliest.strftime('%b')} {earliest.year}"
    return ""


# ── Image fetching (official Meta Graph API) ──────────────────────

def _fetch_fb_image(url: str, fb_token: str, page_id: str) -> bytes | None:
    """Return raw image bytes for a Facebook post.

    Node ID resolution (tested and confirmed):
      /videos/123  /reel/123  /posts/123  → PAGE_ID_123
      /posts/pfbid...  /page/pfbid...     → PAGE_ID_pfbid...  (prepend works directly)
      /photo/?fbid=123                    → PAGE_ID_123

    Fields fetched:
      full_picture                        → single image / video thumbnail
      attachments{media,type,subattachments} → album posts (no top-level full_picture)
    """
    def _download(img_url: str) -> bytes | None:
        try:
            img_r = requests.get(img_url, timeout=15)
            if img_r.status_code == 200:
                raw = img_r.content
                print(f"  [img-fb] Downloaded : {len(raw):,} bytes", end="")
                if _PILImage:
                    try:
                        w, h = _PILImage.open(BytesIO(raw)).size
                        print(f"  →  {w} × {h} px")
                    except Exception:
                        print()
                else:
                    print()
                return raw
        except Exception:
            pass
        return None

    try:
        # ── Resolve node_id from URL ──────────────────────────────────
        node_id = None

        # 1. Numeric ID in path: /videos/123, /reel/123, /posts/123
        m = re.search(r'/(?:posts|videos|reel|reels)/(\d+)', url)
        if m:
            node_id = f"{page_id}_{m.group(1)}"

        # 2. pfbid in path: /posts/pfbid... or /pagename/pfbid...
        #    Confirmed: PAGE_ID_pfbid... works directly as a Graph API node
        if not node_id:
            pfbid_m = re.search(r'/(pfbid[^/?#\s]+)', url)
            if pfbid_m:
                node_id = f"{page_id}_{pfbid_m.group(1)}"
                print(f"  [img-fb] pfbid URL → node: {node_id[:55]}...")

        # 3. Photo URL: /photo/?fbid=123456
        if not node_id:
            fbid_m = re.search(r'[?&]fbid=(\d+)', url)
            if fbid_m:
                node_id = f"{page_id}_{fbid_m.group(1)}"
                print(f"  [img-fb] photo fbid URL → node: {node_id}")

        if not node_id:
            print(f"  [img-fb] Could not extract post ID from URL: {url}")
            return None

        # ── Fetch full_picture + attachments in one call ──────────────
        print(f"  [img-fb] API call: GET /{node_id}?fields=full_picture,attachments{{...}}")
        r = requests.get(
            f"{_BASE_URL}/{node_id}",
            params={
                "fields": "full_picture,attachments{media,type,subattachments}",
                "access_token": fb_token,
            },
            timeout=10,
        )
        data = r.json()

        # Priority 1: top-level full_picture (single photo / video / reel)
        img_url = data.get("full_picture")
        if img_url:
            print(f"  [img-fb] Image URL (full_picture): {img_url}")
            return _download(img_url)

        # Priority 2: album / carousel — no top-level full_picture;
        # use first subattachment image
        for att in data.get("attachments", {}).get("data", []):
            # Direct media on the attachment
            src = att.get("media", {}).get("image", {}).get("src")
            if src:
                print(f"  [img-fb] Image URL (attachment): {src}")
                return _download(src)
            # Subattachments (album)
            for sub in att.get("subattachments", {}).get("data", []):
                src = sub.get("media", {}).get("image", {}).get("src")
                if src:
                    print(f"  [img-fb] Image URL (subattachment[0]): {src}")
                    return _download(src)

        print(f"  [img-fb] No image found in response: {data}")
    except Exception as e:
        print(f"  [img-fb] Error: {e}")
    return None


def _fetch_ig_image(url: str, ig_token: str, ig_user_id: str,
                    ig_media_id: str = "") -> bytes | None:
    """Return raw image bytes for an Instagram post/reel.

    Uses GET /{IG_MEDIA_ID}?fields=media_type,media_url,thumbnail_url,
    children{media_url,media_type} to handle all post types:
      - IMAGE/CAROUSEL_ALBUM → media_url (or first child's media_url)
      - VIDEO (Reel) → thumbnail_url

    Fast path: direct media ID lookup (when ig_media_id available).
    Slow path: search by shortcode through account media list.
    """
    def _download(img_url: str) -> bytes | None:
        if not img_url:
            return None
        print(f"  [img-ig] Image URL: {img_url}")
        try:
            r = requests.get(img_url, timeout=15)
            if r.status_code == 200 and len(r.content) > 100:
                raw = r.content
                print(f"  [img-ig] Downloaded : {len(raw):,} bytes", end="")
                if _PILImage:
                    try:
                        w, h = _PILImage.open(BytesIO(raw)).size
                        print(f"  →  {w} × {h} px")
                    except Exception:
                        print()
                else:
                    print()
                return raw
        except Exception:
            pass
        return None

    def _extract_image_url(data: dict) -> str | None:
        """Pick the best image URL from a media object response."""
        media_type = data.get("media_type", "")
        if media_type == "VIDEO":
            # Reels/videos: use thumbnail
            return data.get("thumbnail_url") or data.get("media_url")
        elif media_type == "CAROUSEL_ALBUM":
            # Carousel: use first child's image
            children = data.get("children", {}).get("data", [])
            for child in children:
                child_type = child.get("media_type", "")
                if child_type in ("IMAGE", ""):
                    return child.get("media_url")
                elif child_type == "VIDEO":
                    return child.get("thumbnail_url") or child.get("media_url")
            # Fallback: carousel's own media_url (sometimes available)
            return data.get("media_url")
        else:
            # IMAGE or unknown: use media_url directly
            return data.get("media_url")

    _FIELDS = "media_type,media_url,thumbnail_url,children{media_url,media_type}"

    try:
        # Fast path: direct media ID lookup
        if ig_media_id and ig_token:
            print(f"  [img] Fetching IG image via media ID: {ig_media_id}")
            r = requests.get(
                f"{_BASE_URL}/{ig_media_id}",
                params={"fields": _FIELDS, "access_token": ig_token},
                timeout=10,
            )
            data = r.json()
            if "error" not in data:
                img_url = _extract_image_url(data)
                result = _download(img_url)
                if result:
                    print(f"  [img] IG image fetched OK ({len(result):,} bytes)")
                    return result
                else:
                    print(f"  [img] IG media_url returned but download failed")
            else:
                print(f"  [img] IG media ID lookup error: {data['error'].get('message', '')}")

        # Slow path: search by shortcode
        m = re.search(r'/(?:p|reel|reels)/([^/?#&]+)', url)
        if not m:
            return None
        shortcode = m.group(1)
        print(f"  [img] Falling back to shortcode search: {shortcode}")

        params = {
            "fields": f"id,shortcode,{_FIELDS}",
            "limit": 100,
            "access_token": ig_token,
        }
        r = requests.get(f"{_BASE_URL}/{ig_user_id}/media", params=params, timeout=15)
        data = r.json()
        item = next((x for x in data.get("data", []) if x.get("shortcode") == shortcode), None)

        while not item and data.get("paging", {}).get("next"):
            r = requests.get(data["paging"]["next"], timeout=15)
            data = r.json()
            item = next((x for x in data.get("data", []) if x.get("shortcode") == shortcode), None)

        if not item:
            print(f"  [img] Shortcode '{shortcode}' not found in media list")
            return None
        img_url = _extract_image_url(item)
        return _download(img_url)
    except Exception as e:
        print(f"  [!] IG image fetch error: {e}")
    return None


def fetch_post_image(
    result: dict,
    fb_token: str = "",
    ig_token: str = "",
    page_id: str = "",
    ig_user_id: str = "",
) -> bytes | None:
    """Dispatch to FB or IG image fetcher based on result platform."""
    fb_token   = fb_token   or os.getenv("FACEBOOK_ACCESS_TOKEN", "")
    ig_token   = ig_token   or os.getenv("IG_ACCESS_TOKEN", "")
    page_id    = page_id    or os.getenv("FACEBOOK_PAGE_ID", "")
    ig_user_id = ig_user_id or os.getenv("IG_BUSINESS_ID", "")

    url      = result.get("url", "")
    platform = result.get("platform", "")

    if platform == "Facebook" or "facebook.com" in url:
        return _fetch_fb_image(url, fb_token, page_id)
    elif platform == "Instagram" or "instagram.com" in url:
        return _fetch_ig_image(url, ig_token, ig_user_id,
                               ig_media_id=result.get("ig_media_id", ""))
    return None


# ── Image sizing helper ───────────────────────────────────────────

def _fill_image(img_bytes: bytes, target_w: int, target_h: int) -> tuple[bytes, int, int]:
    """Resize + center-crop image to fill target_w × target_h exactly.

    Uses scale-to-cover (larger ratio) so no blank space remains, then crops
    to the target rectangle.  Falls back to original bytes if PIL unavailable.
    Returns (processed_bytes, actual_w, actual_h).
    """
    if _PILImage is None:
        return img_bytes, target_w, target_h
    try:
        pil = _PILImage.open(BytesIO(img_bytes)).convert("RGB")
        orig_w, orig_h = pil.size
        ratio = max(target_w / orig_w, target_h / orig_h)
        new_w, new_h = int(orig_w * ratio), int(orig_h * ratio)
        try:
            resample = _PILImage.Resampling.LANCZOS
        except AttributeError:
            resample = _PILImage.LANCZOS
        pil = pil.resize((new_w, new_h), resample)
        left = (new_w - target_w) // 2
        top  = (new_h - target_h) // 2
        pil  = pil.crop((left, top, left + target_w, top + target_h))
        out  = BytesIO()
        pil.save(out, format="PNG")
        out.seek(0)
        return out.read(), target_w, target_h
    except Exception:
        return img_bytes, target_w, target_h


def _fit_image(img_bytes: bytes, max_w: int, max_h: int) -> tuple[bytes, int, int]:
    """Resize image to fit within max_w × max_h, preserving aspect ratio.

    Scales down proportionally so the full image is visible (no cropping).
    Uses high-quality LANCZOS resampling for a clean result.
    Returns (processed_bytes, actual_w, actual_h).
    """
    if _PILImage is None:
        return img_bytes, max_w, max_h
    try:
        pil = _PILImage.open(BytesIO(img_bytes)).convert("RGB")
        orig_w, orig_h = pil.size
        scale = min(max_w / orig_w, max_h / orig_h)
        new_w, new_h = int(orig_w * scale), int(orig_h * scale)
        try:
            resample = _PILImage.Resampling.LANCZOS
        except AttributeError:
            resample = _PILImage.LANCZOS
        pil = pil.resize((new_w, new_h), resample)
        out = BytesIO()
        pil.save(out, format="PNG")
        out.seek(0)
        return out.read(), new_w, new_h
    except Exception:
        return img_bytes, max_w, max_h


# ── Cell writers ──────────────────────────────────────────────────

def _write_section_title(ws, row: int, text: str):
    """Write section title (e.g. 'Post Insight', 'Ad Boosting Report') — no fill, black bold."""
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = _font(bold=True, color=_BLACK, size=14)
    cell.alignment = _align("left", wrap=False)


def _write_label(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = _font(bold=True, size=14)
    cell.alignment = _align("left", wrap=False)


def _write_value(ws, row: int, col: int, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = _font(size=14)
    cell.alignment = _align("left", wrap=False)


def _write_data(ws, row: int, col: int, value, bold=False, color=_BLACK,
                halign="center", underline=False, num_fmt=None, wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, color=color, size=14, underline=underline)
    cell.border = _border()
    cell.alignment = _align(halign, wrap=wrap)
    # Auto thousand separator for integers
    if num_fmt:
        cell.number_format = num_fmt
    elif isinstance(value, int) and value != 0:
        cell.number_format = "#,##0"
    return cell


# ── Main generator ────────────────────────────────────────────────

def generate_excel_report(
    result: dict,
    fb_token: str = "",
    ig_token: str = "",
    page_id: str = "",
    ig_user_id: str = "",
) -> BytesIO:
    """
    Build a Workbook matching the client post-buy report template.
    Returns a BytesIO buffer (no temp files created).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Post Report"
    ws.sheet_view.showGridLines = False

    # Column widths
    for col_num, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Row heights
    ws.row_dimensions[1].height  = 6
    for r in range(2, 8):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[8].height  = 6
    ws.row_dimensions[9].height  = 368.5  # image (~360pt)
    ws.row_dimensions[10].height = 8
    ws.row_dimensions[11].height = 22   # orange header
    ws.row_dimensions[12].height = 16
    ws.row_dimensions[13].height = 42   # column headers
    for r in range(14, 19):
        ws.row_dimensions[r].height = 36
    ws.row_dimensions[19].height = 30   # remark
    ws.row_dimensions[20].height = 28   # ≈ 1 cm separator
    ws.row_dimensions[21].height = 22   # "Ad Boosting Report" title
    ws.row_dimensions[22].height = 16   # ≈ 0.56 cm gap below title
    ws.row_dimensions[23].height = 42   # ad column headers

    # ── Parse info ────────────────────────────────────────────────
    ad_list_raw = result.get("ad_metrics", [])
    adset_names    = [am.get("adset_name", "") for am in ad_list_raw]
    campaign_names = [am.get("campaign_name", "") for am in ad_list_raw]
    campaign    = parse_campaign_header(adset_names, campaign_names)
    launch_date = _get_ad_start_date(ad_list_raw)
    platform    = result.get("platform", "Unknown")
    url         = result.get("url", "")
    is_ig       = platform == "Instagram" or "instagram.com" in url

    # ── Section 1: Campaign Header ────────────────────────────────
    _write_label(ws, 2, 2, "Campaign:")
    campaign_display = f"{campaign['name']} Post-buy Report" if campaign["name"] else ""
    _write_value(ws, 2, 3, campaign_display)

    _write_label(ws, 3, 2, "Code:")
    code_display = f"#{campaign['code']}" if campaign["code"] else ""
    _write_value(ws, 3, 3, code_display)

    _write_label(ws, 4, 2, "Launch Date:")
    _write_value(ws, 4, 3, launch_date)

    _write_label(ws, 5, 2, "Placement:")
    if is_ig:
        _write_value(ws, 5, 3, "Instagram Feed")
    else:
        _write_value(ws, 5, 3, "Facebook Feed")

    # ── Section 2: Post Image ─────────────────────────────────────
    img_bytes = fetch_post_image(
        result,
        fb_token=fb_token,
        ig_token=ig_token,
        page_id=page_id,
        ig_user_id=ig_user_id,
    )
    if img_bytes:
        try:
            fitted, iw, ih = _fit_image(img_bytes, 369, 491)
            xl_img = XLImage(BytesIO(fitted))
            xl_img.width, xl_img.height = iw, ih
            ws.add_image(xl_img, "C9")
        except Exception:   
            ws.cell(row=9, column=3, value="[Image could not be embedded]")
    else:
        cell = ws.cell(row=9, column=3, value="[Paste screenshot here]")
        cell.font = _font(color=_GREY_TXT, italic=True)
        cell.alignment = _align("center")

    # ── Section 3: Post Insight Table ─────────────────────────────
    _write_section_title(ws, 11, "Post Insight  (Overall Performance)")
    for col in range(1, 12):
        ws.cell(row=11, column=col).border = Border(bottom=Side(style="medium", color="5B9BD5"))
    # Column headers row 13
    post_cols = [
        "Platform", "Post Link", "Views", "Reach", "Interactions",
        "Likes and reactions", "Comments", "Shares", "Saves", "Link clicks",
    ]
    for ci, h in enumerate(post_cols, start=2):
        c = _write_data(ws, 13, ci, h, bold=True, halign="center")
        c.fill = _fill(_ORANGE)
        c.font = _font(bold=True, color=_WHITE, size=14)

    # Metric values
    views        = result.get("views", 0)
    reach        = result.get("reach", 0)
    interactions = result.get("interactions", 0)
    reactions    = result.get("reactions", 0)
    comments     = result.get("comments", 0)
    shares       = result.get("shares", 0)
    saves        = result.get("saves", 0)
    link_clicks  = result.get("link_clicks", 0)

    # Row layout: FB Post (14), FB Story (15), IG Post (16), IG Story (17)
    rows_def = [
        (14, "FB Post",  False),
        (15, "FB Story", False),
        (16, "IG Post",  False),
        (17, "IG Story", False),
    ]
    if is_ig:
        rows_def[2] = (16, "IG Post", True)   # IG Post filled
    else:
        rows_def[0] = (14, "FB Post", True)   # FB Post filled

    for (row_num, plat_label, filled) in rows_def:
        _write_data(ws, row_num, 2, plat_label, halign="left")
        # URL cell
        url_val = url if filled else ""
        url_cell = _write_data(
            ws, row_num, 3, url_val,
            color=_BLUE if url_val else _BLACK,
            underline=bool(url_val),
            halign="left",
        )
        if url_val:
            url_cell.hyperlink = url_val

        metrics = [views, reach, interactions, reactions, comments, shares, saves, link_clicks] if filled else [0]*8
        for ci, val in enumerate(metrics, start=4):
            _write_data(ws, row_num, ci, val if (filled and val) else "", halign="center")

    # Total row 18
    _write_data(ws, 18, 2, "Total", halign="left")
    _write_data(ws, 18, 3, "")
    for ci, val in enumerate([views, reach, interactions, reactions, comments, shares, saves, link_clicks], start=4):
        _write_data(ws, 18, ci, val, halign="center")

    # Remark row 19
    remark1 = (
        "Remark: Paid result of insight report maybe various in boosting report, "
        "based on different metric counts in Post Insight and Ads Manager."
    )
    r1 = ws.cell(row=19, column=2, value=remark1)
    r1.font      = _font(size=14, color=_BLACK)
    r1.alignment = _align("left", wrap=False)
 
    # ── Section 4: Ad Boosting Report ─────────────────────────────
    _write_section_title(ws, 21, "Ad Boosting Report")
    for col in range(1, 12):
        ws.cell(row=21, column=col).border = Border(bottom=Side(style="medium", color="5B9BD5"))

    # Column headers row 23 — "Ad set name" spans merged B+C; data cols start at D (4)
    ws.merge_cells("B23:C23")
    c = _write_data(ws, 23, 2, "Ad set name", bold=True, halign="center")
    c.fill = _fill(_ORANGE)
    c.font = _font(bold=True, color=_WHITE, size=14)
    for ci, h in enumerate(
        ["Amount spent (HKD)", "Impressions", "Link Clicks",
         "Clicks (All)", "Post engagements", "Post reactions", "Post comments",
         "Reach", "Post shares", "Post saves"],
        start=4,
    ):
        c = _write_data(ws, 23, ci, h, bold=True, halign="center")
        c.fill = _fill(_ORANGE)
        c.font = _font(bold=True, color=_WHITE, size=14)

    # Ad data rows — adset name occupies merged B+C; other metrics from col D onward
    ad_list  = result.get("ad_metrics", [])
    last_row = 24
    if ad_list:
        for i, am in enumerate(ad_list):
            r = 24 + i
            ws.row_dimensions[r].height = 36
            ws.merge_cells(f"B{r}:C{r}")
            _write_data(ws, r, 2, am.get("adset_name", "N/A"), halign="left", wrap=False)
            _write_data(ws, r, 4, round(am.get("spend", 0), 2),
                        halign="center", num_fmt="#,##0.00")
            _write_data(ws, r, 5, am.get("impressions", 0),    halign="center")
            _write_data(ws, r, 6, am.get("link_clicks", 0),    halign="center")
            _write_data(ws, r, 7, am.get("clicks_all", 0),     halign="center")
            _write_data(ws, r, 8, am.get("post_engagement", 0),halign="center")
            _write_data(ws, r, 9, am.get("reactions", 0),      halign="center")
            _write_data(ws, r, 10, am.get("comments", 0),      halign="center")
            _write_data(ws, r, 11, am.get("reach", 0),         halign="center")
            _write_data(ws, r, 12, am.get("shares", 0),        halign="center")
            _write_data(ws, r, 13, am.get("saves", 0),         halign="center")
        # Total row
        total_row = 24 + len(ad_list)
        ws.row_dimensions[total_row].height = 36
        ws.merge_cells(f"B{total_row}:C{total_row}")
        _write_data(ws, total_row, 2, "Total", halign="left")
        _write_data(ws, total_row, 4, round(sum(am.get("spend", 0) for am in ad_list), 2),
                    halign="center", num_fmt="#,##0.00")
        _write_data(ws, total_row, 5, sum(am.get("impressions", 0) for am in ad_list),
                    halign="center")
        _write_data(ws, total_row, 6, sum(am.get("link_clicks", 0) for am in ad_list),
                    halign="center")
        _write_data(ws, total_row, 7, sum(am.get("clicks_all", 0) for am in ad_list),
                    halign="center")
        _write_data(ws, total_row, 8, sum(am.get("post_engagement", 0) for am in ad_list),
                    halign="center")
        _write_data(ws, total_row, 9, sum(am.get("reactions", 0) for am in ad_list),
                    halign="center")
        _write_data(ws, total_row, 10, sum(am.get("comments", 0) for am in ad_list),
                    halign="center")
        _write_data(ws, total_row, 11, sum(am.get("reach", 0) for am in ad_list),
                    halign="center")
        _write_data(ws, total_row, 12, sum(am.get("shares", 0) for am in ad_list),
                    halign="center")
        _write_data(ws, total_row, 13, sum(am.get("saves", 0) for am in ad_list),
                    halign="center")
        last_row = total_row + 1
    else:
        ws.merge_cells("B24:C24")
        _write_data(ws, 24, 2, "No ad data", halign="left")
        last_row = 25

    # Final remark
    remark2 = (
        "Remark: Manually accumulated Reach or other figures from individual Ad Sets "
        "may vary from Total as stated from the report due to different counting "
        "methodologies/sampling methods by Facebook."
    )
    r2 = ws.cell(row=last_row, column=2, value=remark2)
    r2.font      = _font(size=14, color=_BLACK)
    r2.alignment = _align("left", wrap=False)

    # ── Write to buffer ───────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Combined FB + IG report ───────────────────────────────────────

def generate_excel_report_combined(
    results: list,
    fb_token: str = "",
    ig_token: str = "",
    page_id: str = "",
    ig_user_id: str = "",
) -> BytesIO:
    """
    Build a combined Workbook for a set of results (typically one FB + one IG).
    Rows 14/15 = FB Post / FB Story, rows 16/17 = IG Post / IG Story.
    Images: FB post at C9, FB story placeholder at F9, IG post at H9, IG story placeholder at K9.
    Ad metrics from all results are combined in the Ad Boosting Report section.
    """
    fb_token   = fb_token   or os.getenv("FACEBOOK_ACCESS_TOKEN", "")
    ig_token   = ig_token   or os.getenv("IG_ACCESS_TOKEN", "")
    page_id    = page_id    or os.getenv("FACEBOOK_PAGE_ID", "")
    ig_user_id = ig_user_id or os.getenv("IG_BUSINESS_ID", "")

    fb_result = next(
        (r for r in results if r.get("platform") == "Facebook" or "facebook.com" in r.get("url", "")),
        None,
    )
    ig_result = next(
        (r for r in results if r.get("platform") == "Instagram" or "instagram.com" in r.get("url", "")),
        None,
    )
    primary = fb_result or ig_result
    if not primary:
        raise ValueError("No results to generate Excel report from.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Post Report"
    ws.sheet_view.showGridLines = False

    for col_num, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.row_dimensions[1].height  = 6
    for r in range(2, 8):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[8].height  = 6
    ws.row_dimensions[9].height  = 368.5
    ws.row_dimensions[10].height = 8
    ws.row_dimensions[11].height = 22
    ws.row_dimensions[12].height = 16
    ws.row_dimensions[13].height = 42
    for r in range(14, 19):
        ws.row_dimensions[r].height = 36
    ws.row_dimensions[19].height = 30
    ws.row_dimensions[20].height = 28   # ≈ 1 cm separator
    ws.row_dimensions[21].height = 22   # "Ad Boosting Report" title
    ws.row_dimensions[22].height = 16   # ≈ 0.56 cm gap below title
    ws.row_dimensions[23].height = 42   # ad column headers

    # ── Campaign header ───────────────────────────────────────────
    adset_names = []
    campaign_names = []
    for res in results:
        adset_names += [am.get("adset_name", "") for am in res.get("ad_metrics", [])]
        campaign_names += [am.get("campaign_name", "") for am in res.get("ad_metrics", [])]
    campaign = parse_campaign_header(adset_names, campaign_names)
    all_ad_metrics = []
    for res in results:
        all_ad_metrics += res.get("ad_metrics", [])
    launch_date = _get_ad_start_date(all_ad_metrics)

    _write_label(ws, 2, 2, "Campaign:")
    campaign_display = f"{campaign['name']} Post-buy Report" if campaign["name"] else ""
    _write_value(ws, 2, 3, campaign_display)
    _write_label(ws, 3, 2, "Code:")
    code_display = f"#{campaign['code']}" if campaign["code"] else ""
    _write_value(ws, 3, 3, code_display)
    _write_label(ws, 4, 2, "Launch Date:")
    _write_value(ws, 4, 3, launch_date)
    _write_label(ws, 5, 2, "Placement:")
    # Placement based on which platforms are present
    has_fb = any(r.get("platform") == "Facebook" or "facebook.com" in r.get("url", "") for r in results)
    has_ig = any(r.get("platform") == "Instagram" or "instagram.com" in r.get("url", "") for r in results)
    if has_fb and has_ig:
        _write_value(ws, 5, 3, "Facebook and Instagram Feed")
    elif has_fb:
        _write_value(ws, 5, 3, "Facebook Feed")
    elif has_ig:
        _write_value(ws, 5, 3, "Instagram Feed")

    # ── Images: FB post (C9), FB Story placeholder (F9),
    #           IG post (H9), IG Story placeholder (K9) ────────────
    def _story_placeholder(col, label):
        cell = ws.cell(row=9, column=col, value=f"{label}\n[Paste screenshot]")
        cell.font      = _font(color=_GREY_TXT, italic=True, size=10)
        cell.alignment = _align("center")

    if fb_result:
        fb_img = fetch_post_image(fb_result, fb_token=fb_token, ig_token=ig_token,
                                  page_id=page_id, ig_user_id=ig_user_id)
        if fb_img:
            try:
                fitted, iw, ih = _fit_image(fb_img, 369, 491)
                xl = XLImage(BytesIO(fitted))
                xl.width, xl.height = iw, ih
                ws.add_image(xl, "C9")
            except Exception:
                ws.cell(row=9, column=3, value="[FB image unavailable]")
        else:
            ws.cell(row=9, column=3, value="[Paste FB post screenshot]")
    _story_placeholder(6, "Facebook Story")   # col F

    if ig_result:
        ig_img = fetch_post_image(ig_result, fb_token=fb_token, ig_token=ig_token,
                                  page_id=page_id, ig_user_id=ig_user_id)
        if ig_img:
            try:
                fitted, iw, ih = _fit_image(ig_img, 369, 491)
                xl = XLImage(BytesIO(fitted))
                xl.width, xl.height = iw, ih
                ws.add_image(xl, "H9")
            except Exception:
                ws.cell(row=9, column=8, value="[IG image unavailable]")
        else:
            ws.cell(row=9, column=8, value="[Paste IG post screenshot]")
    _story_placeholder(11, "Instagram Story")  # col K

    # ── Post Insight table ────────────────────────────────────────
    _write_section_title(ws, 11, "Post Insight  (Overall Performance)")
    for col in range(1, 12):
        ws.cell(row=11, column=col).border = Border(bottom=Side(style="medium", color="5B9BD5"))
    post_cols = [
        "Platform", "Post Link", "Views", "Reach", "Interactions",
        "Likes and reactions", "Comments", "Shares", "Saves", "Link clicks",
    ]
    for ci, h in enumerate(post_cols, start=2):
        c = _write_data(ws, 13, ci, h, bold=True, halign="center")
        c.fill = _fill(_ORANGE)
        c.font = _font(bold=True, color=_WHITE, size=14)

    def _metrics(res):
        if not res:
            return [0] * 8
        return [
            res.get("views", 0), res.get("reach", 0), res.get("interactions", 0),
            res.get("reactions", 0), res.get("comments", 0), res.get("shares", 0),
            res.get("saves", 0), res.get("link_clicks", 0),
        ]

    fb_m = _metrics(fb_result)
    ig_m = _metrics(ig_result)

    for row_num, label, res, m in [
        (14, "FB Post",   fb_result, fb_m),
        (15, "FB Story",  None,      [0] * 8),
        (16, "IG Post",   ig_result, ig_m),
        (17, "IG Story",  None,      [0] * 8),
    ]:
        _write_data(ws, row_num, 2, label, halign="left")
        url_val = res.get("url", "") if res else ""
        url_cell = _write_data(ws, row_num, 3, url_val,
                               color=_BLUE if url_val else _BLACK,
                               underline=bool(url_val), halign="left")
        if url_val:
            url_cell.hyperlink = url_val
        for ci, val in enumerate(m, start=4):
            _write_data(ws, row_num, ci, val if (res and val) else "", halign="center")

    # Total row
    total = [a + b for a, b in zip(fb_m, ig_m)]
    _write_data(ws, 18, 2, "Total", halign="left")
    _write_data(ws, 18, 3, "")
    for ci, val in enumerate(total, start=4):
        _write_data(ws, 18, ci, val, halign="center")

    remark1 = (
        "Remark: Paid result of insight report maybe various in boosting report, "
        "based on different metric counts in Post Insight and Ads Manager."
    )
    r1 = ws.cell(row=19, column=2, value=remark1)
    r1.font      = _font(size=14, color=_BLACK)
    r1.alignment = _align("left", wrap=False)

    # ── Ad Boosting Report ────────────────────────────────────────
    _write_section_title(ws, 21, "Ad Boosting Report")
    for col in range(1, 12):
        ws.cell(row=21, column=col).border = Border(bottom=Side(style="medium", color="5B9BD5"))

    # Column headers row 23 — "Ad set name" spans merged B+C; data cols start at D (4)
    ws.merge_cells("B23:C23")
    c = _write_data(ws, 23, 2, "Ad set name", bold=True, halign="center")
    c.fill = _fill(_ORANGE)
    c.font = _font(bold=True, color=_WHITE, size=14)
    for ci, h in enumerate(
        ["Amount spent (HKD)", "Impressions", "Link Clicks",
         "Clicks (All)", "Post engagements", "Post reactions", "Post comments",
         "Reach", "Post shares", "Post saves"],
        start=4,
    ):
        c = _write_data(ws, 23, ci, h, bold=True, halign="center")
        c.fill = _fill(_ORANGE)
        c.font = _font(bold=True, color=_WHITE, size=14)

    all_ads = []
    for res in results:
        all_ads += res.get("ad_metrics", [])

    # Ad data rows — adset name occupies merged B+C; other metrics from col D onward
    last_row = 24
    if all_ads:
        for idx, am in enumerate(all_ads):
            r = 24 + idx
            ws.row_dimensions[r].height = 36
            ws.merge_cells(f"B{r}:C{r}")
            _write_data(ws, r, 2, am.get("adset_name", "N/A"), halign="left", wrap=False)
            _write_data(ws, r, 4, round(am.get("spend", 0), 2),        halign="center", num_fmt="#,##0.00")
            _write_data(ws, r, 5, am.get("impressions", 0),            halign="center")
            _write_data(ws, r, 6, am.get("link_clicks", 0),            halign="center")
            _write_data(ws, r, 7, am.get("clicks_all", 0),             halign="center")
            _write_data(ws, r, 8, am.get("post_engagement", 0),        halign="center")
            _write_data(ws, r, 9, am.get("reactions", 0),              halign="center")
            _write_data(ws, r, 10, am.get("comments", 0),              halign="center")
            _write_data(ws, r, 11, am.get("reach", 0),                 halign="center")
            _write_data(ws, r, 12, am.get("shares", 0),                halign="center")
            _write_data(ws, r, 13, am.get("saves", 0),                 halign="center")
        # Total row
        total_row = 24 + len(all_ads)
        ws.row_dimensions[total_row].height = 36
        ws.merge_cells(f"B{total_row}:C{total_row}")
        _write_data(ws, total_row, 2, "Total", halign="left")
        _write_data(ws, total_row, 4, round(sum(am.get("spend", 0) for am in all_ads), 2),
                    halign="center", num_fmt="#,##0.00")
        _write_data(ws, total_row, 5, sum(am.get("impressions", 0) for am in all_ads),
                    halign="center")
        _write_data(ws, total_row, 6, sum(am.get("link_clicks", 0) for am in all_ads),
                    halign="center")
        _write_data(ws, total_row, 7, sum(am.get("clicks_all", 0) for am in all_ads),
                    halign="center")
        _write_data(ws, total_row, 8, sum(am.get("post_engagement", 0) for am in all_ads),
                    halign="center")
        _write_data(ws, total_row, 9, sum(am.get("reactions", 0) for am in all_ads),
                    halign="center")
        _write_data(ws, total_row, 10, sum(am.get("comments", 0) for am in all_ads),
                    halign="center")
        _write_data(ws, total_row, 11, sum(am.get("reach", 0) for am in all_ads),
                    halign="center")
        _write_data(ws, total_row, 12, sum(am.get("shares", 0) for am in all_ads),
                    halign="center")
        _write_data(ws, total_row, 13, sum(am.get("saves", 0) for am in all_ads),
                    halign="center")
        last_row = total_row + 1
    else:
        ws.merge_cells("B24:C24")
        _write_data(ws, 24, 2, "No ad data", halign="left")
        last_row = 25

    remark2 = (
        "Remark: Manually accumulated Reach or other figures from individual Ad Sets "
        "may vary from Total as stated from the report due to different counting "
        "methodologies/sampling methods by Facebook."
    )
    r2 = ws.cell(row=last_row, column=2, value=remark2)
    r2.font      = _font(size=14, color=_BLACK)
    r2.alignment = _align("left", wrap=False)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── CLI quick-test ────────────────────────────────────────────────
if __name__ == "__main__":
    dummy = {
        "platform": "Facebook",
        "url": "https://www.facebook.com/playeateasy/posts/1192897706280130",
        "date_range": "2026-01-08 to 2026-01-12",
        "views": 109174,
        "reach": 102673,
        "interactions": 474,
        "reactions": 455,
        "comments": 5,
        "shares": 11,
        "saves": 3,
        "link_clicks": 8,
        "ad_metrics": [
            {
                "adset_name": "FB (Reach) (Bank) #172072 DBS - Year End Promotion 2025",
                "spend": 1500.00,
                "impressions": 80000,
                "link_clicks": 320,
                "clicks_all": 450,
                "post_engagement": 890,
                "reactions": 455,
                "comments": 5,
                "shares": 0,
            }
        ],
    }
    buf = generate_excel_report(dummy)
    with open("test_report.xlsx", "wb") as f:
        f.write(buf.read())
    print("Written: test_report.xlsx — open in Excel to verify.")
